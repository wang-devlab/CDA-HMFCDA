from gensim.models import KeyedVectors
from openpyxl import load_workbook
from tqdm import tqdm  # 进度条工具
import numpy as np
import scipy.io as sio  # 新增用于.mat保存的库

# biowordvec 疾病名称转为向量

# 1. 加载BioWordVec模型（需提前下载biowordvec.bin）
print("正在加载BioWordVec模型...")
# 指定模型路径
model_path = "./bioword2vec.bin"
# 加载模型（二进制格式）
model = KeyedVectors.load_word2vec_format(model_path, binary=True)
print(f"模型加载完成，包含 {len(model.key_to_index)} 个词向量")

# 2. 疾病名称转向量函数
def disease_to_vector(disease_name, model):
    """基础处理：小写化、空格分割"""
    words = str(disease_name).lower().split()
    vectors = [model[word] for word in words if word in model]
    return np.mean(vectors, axis=0) if vectors else np.zeros(model.vector_size)

# 3. 读取Excel文件
excel_path = "CircR2Disease_Association_Matrixs.xlsx"
sheet_name = "Disease Names"

print(f"\n正在读取 {excel_path} 的 {sheet_name}...")
wb = load_workbook(excel_path)
sheet = wb[sheet_name]

# 4. 处理数据（假设疾病在第一列）
vectors = []
disease_names = []

for row in tqdm(sheet.iter_rows(min_row=1, values_only=True), desc="处理中"):
    if not row[0]:  # 跳过空行
        continue
    
    disease_name = str(row[0]).strip()
    vector = disease_to_vector(disease_name, model)
    
    disease_names.append(disease_name)
    vectors.append(vector)

# 5. 保存为MATLAB格式
mat_data = {
    'disease_names': np.array(disease_names, dtype=object).reshape(-1, 1),  # 转为列向量
    'disease_vectors': np.array(vectors),
    'disease_vector_dim': model.vector_size
}

mat_path = "disease_vectors200.mat"
sio.savemat(mat_path, mat_data)
print(f"\nMATLAB格式已保存到 {mat_path}")

# 6. 可选：同时保存Excel（原功能保留）
output_path = "disease_vectors200.xlsx"
output_wb = load_workbook(excel_path)
if "向量结果" in output_wb.sheetnames:
    output_wb.remove(output_wb["向量结果"])
output_ws = output_wb.create_sheet("向量结果")

output_ws.append(["疾病名称", "向量维度", "向量数据"])
for name, vec in zip(disease_names, vectors):
    output_ws.append([name, len(vec), str(vec.tolist())])

output_wb.save(output_path)
print(f"Excel格式已保存到 {output_path}")

# 7. 打印统计信息
print("\n处理统计：")
print(f"- 总处理疾病数：{len(vectors)}")
print(f"- 向量维度：{model.vector_size}")
print("- MATLAB文件内容：")
print(f"  |-- disease_names: {len(disease_names)}x1 cell array")
print(f"  |-- vectors: {len(vectors)}x{model.vector_size} matrix")
print(f"  |-- vector_dim: scalar value")